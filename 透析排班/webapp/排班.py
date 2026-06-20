# -*- coding: utf-8 -*-
"""
透析「印藥水(列印備物清單)」每週自動排班程式  v0.4
===============================================================
v0.4 重點(納入組長 2026-06-10 回覆):
  • EP1☆ 改為不可印(可能被通知放假);病房/ICU/大夜/長假「都算休息」進公平輪序;
    跨區從「組員較多的區」借人;印兩次盡量間隔、同日一+二也允許。
v0.3 重點:
  • 採用已驗證的正確邏輯:名單按「治療日」、白班印前1個治療日、小夜印前2個治療日,
    跨週末/連假會跳過(不是死板的日曆減1/2);名單窗口=星期二~下週一。
  • 床號分區、組員、不可印班別、休診日(連假)全部「設定檔化」,可用 Excel 編輯,
    程式啟動會自動建立預設檔。要開「第三洗腎室」只要在床號檔加『三區』的列即可。
  • 全面防呆:未知床號會警告(不默默漏人)、排不出的格子留白標❌、連假自動跳過、
    卡號變更用姓名補抓、重複/缺檔/壞資料都給友善訊息不當機。

用法:  python 排班.py <班表.xls> [分頁名稱]
        分頁不填 → 自動用最後一個分頁(最新一週)
"""
import os, sys, csv, traceback
from datetime import date, datetime, timedelta
import pandas as pd

# ===================== 路徑 =====================
BASE = os.path.dirname(os.path.abspath(__file__))
F_ROSTER  = os.path.join(BASE, "組員名單.csv")
F_BEDS    = os.path.join(BASE, "床號分區.csv")
F_SHIFTS  = os.path.join(BASE, "不可印班別.csv")
F_HOLIDAY = os.path.join(BASE, "休診日.csv")
F_HISTORY = os.path.join(BASE, "排班紀錄.csv")
F_README  = os.path.join(BASE, "設定說明.txt")
OUT_DIR   = os.path.join(BASE, "輸出")
os.makedirs(OUT_DIR, exist_ok=True)

WARN = []                       # 警告訊息
def warn(msg): WARN.append(msg)

# ===================== 預設設定(第一次跑自動建立) =====================
DEF_ROSTER = [
    ("5379","巫佳容"),("G343","高翠盈"),("6663","劉于甄"),("6344","林欣儀"),
    ("A029","于淑鳳"),("G461","陳葒萍"),("A019","顏凰任"),("4710","廖緗玗"),
    ("4162","張雅雯"),("4890","潘冠秀"),("G645","黃怡璇"),("4825","王乃君"),
    ("4243","黃莉婷"),("6412","余伶緣"),("4281","邱玉繡"),
]
# (區域類型, 床號, 盡量避開Y/N)   區域類型=一區/二區/三區… 或 排除/勤務
DEF_BEDS = [
    ("一區","1-3,21","N"),("一區","5-8","N"),("一區","9-12","N"),("一區","13-17","N"),
    ("一區","18-20","N"),("一區","22-23","Y"),("一區","22-2329-30","Y"),("一區","29-30","Y"),
    ("一區","25-28","N"),("一區","35-59","N"),
    ("二區","60-63","N"),("二區","65-68","N"),("二區","71-75","N"),("二區","76-79","N"),
    ("二區","80-83","N"),("二區","85-89","N"),("二區","85-88","N"),
    ("排除","31-33","N"),("排除","37-39","N"),("排除","37","N"),("排除","50-52","N"),("排除","53-56","N"),
    ("勤務","收水","N"),("勤務","對帳","N"),("勤務","文書","N"),("勤務","查帳","N"),("勤務","W103","N"),
]
DEF_SHIFTS = ["N","11-7","2-10","3'-12","1'-10","12-8","EP1",
              "副值","收水","對帳","文書","查帳","W103","1線","2線","3線","4線","8-5'","8-4'","1-9"]
README = """透析印藥水排班 — 設定檔說明(都可以用 Excel 打開編輯,改完存檔即可)
=====================================================================
1) 組員名單.csv      欄位:卡號, 姓名
   • 增減組員:直接加一列 / 刪一列。卡號不確定可只填姓名,程式會用姓名找。
   • 卡號轉正職變號了沒關係,程式會用姓名自動抓到並提醒你更新。

2) 床號分區.csv      欄位:區域類型, 床號, 盡量避開
   • 區域類型填「一區」「二區」…要開【第三洗腎室】就新增幾列「三區」+ 它的床號,
     程式會自動多排一個三區,不用改程式。
   • 病房/ICU 等不印的床號填「排除」;非床邊勤務填「勤務」。
   • 「盡量避開」填 Y 的床號(例如 22-23),程式會盡量不排他印,除非那天排不出別人。
   • 床號要跟班表上「責任區域」欄寫的『一字不差』(例如 1-3,21 / 22-2329-30)。

3) 不可印班別.csv    欄位:班別
   • 這些班別的人當天不能印(大夜 N、11-7、2-10'、3'-12、1'-10、副值…)。要加就加一列。

4) 休診日.csv        欄位:日期, 說明
   • 過年/國定假日/颱風等「沒有透析」的日子,填進來(日期格式 2026-06-15),
     程式排班會自動跳過那天、提前印的天數也會自動順延略過。

5) 排班紀錄.csv      程式自動維護(記每週誰印誰休,用來算長期公平),不用手動改。
"""

# ===================== 小工具:安全建立 / 讀取 CSV =====================
def _write_csv(path, header, rows):
    with open(path,"w",newline="",encoding="utf-8-sig") as f:
        w=csv.writer(f); w.writerow(header)
        for r in rows: w.writerow(r)

def ensure_configs():
    if not os.path.exists(F_ROSTER):
        _write_csv(F_ROSTER, ["卡號","姓名"], DEF_ROSTER)
        warn(f"※ 已建立預設【組員名單】:{F_ROSTER}")
    if not os.path.exists(F_BEDS):
        _write_csv(F_BEDS, ["區域類型","床號","盡量避開"], DEF_BEDS)
        warn(f"※ 已建立預設【床號分區】:{F_BEDS}")
    if not os.path.exists(F_SHIFTS):
        _write_csv(F_SHIFTS, ["班別"], [[s] for s in DEF_SHIFTS])
        warn(f"※ 已建立預設【不可印班別】:{F_SHIFTS}")
    if not os.path.exists(F_HOLIDAY):
        _write_csv(F_HOLIDAY, ["日期","說明"], [])
        warn(f"※ 已建立空的【休診日】:{F_HOLIDAY}(過年/國定假日可填這裡)")
    if not os.path.exists(F_README):
        with open(F_README,"w",encoding="utf-8") as f: f.write(README)

def _read_rows(path):
    if not os.path.exists(path):       # 檔案還沒建立=正常,不用警告
        return []
    try:
        with open(path, encoding="utf-8-sig") as f:
            return list(csv.DictReader(f))
    except Exception as e:
        warn(f"⚠ 讀取 {os.path.basename(path)} 失敗({e}),改用預設值"); return []

# ===================== 載入設定 =====================
def load_beds():
    rows=_read_rows(F_BEDS); zone2area={}; avoid=set(); order=[]
    for r in rows:
        area=(r.get("區域類型") or "").strip(); bed=(r.get("床號") or "").strip()
        if not area or not bed: continue
        zone2area[bed]=area
        if (r.get("盡量避開") or "").strip().upper()=="Y": avoid.add(bed)
        if area not in ("排除","勤務") and area not in order: order.append(area)
    if not zone2area:    # 壞檔保底
        for a,b,av in DEF_BEDS:
            zone2area[b]=a
            if av=="Y": avoid.add(b)
            if a not in ("排除","勤務") and a not in order: order.append(a)
    return zone2area, avoid, order

def load_shifts():
    rows=_read_rows(F_SHIFTS); lst=[(r.get("班別") or "").strip() for r in rows]
    lst=[s for s in lst if s]
    return lst or list(DEF_SHIFTS)

def load_holidays():
    rows=_read_rows(F_HOLIDAY); hs=set()
    for r in rows:
        s=(r.get("日期") or "").strip()
        if not s: continue
        try: hs.add(datetime.strptime(s[:10], "%Y-%m-%d").date())
        except Exception: warn(f"⚠ 休診日『{s}』格式不對(要 2026-06-15 這種),已略過")
    return hs

def load_roster():
    rows=_read_rows(F_ROSTER); seen=set(); roster=[]
    for r in rows:
        c=(r.get("卡號") or "").strip(); n=(r.get("姓名") or "").strip()
        if not c and not n: continue
        key=(c,n)
        if key in seen: warn(f"⚠ 名單有重複:{n}({c}),只算一次"); continue
        seen.add(key); roster.append({"card":c,"name":n})
    if not roster: warn("❌ 組員名單是空的!請在 組員名單.csv 填入組員")
    return roster

def save_roster(roster):
    _write_csv(F_ROSTER, ["卡號","姓名"], [(m["card"],m["name"]) for m in roster])

# ===================== 全域設定容器 =====================
ZONE2AREA={}; AVOID=set(); PRINT_AREAS=[]; SHIFT_EXCLUDE=[]; HOLIDAYS=set()
UNKNOWN_BEDS={}     # 床號字串 -> set((姓名,日期))

def is_excluded_shift(s):
    s=(s or "").strip()
    if not s: return False
    for tok in SHIFT_EXCLUDE:
        if s==tok: return True
        if len(tok)>=3 and tok in s: return True
    return False

def zone_area(zone):
    z=(zone or "").strip()
    if z=="": return "空"
    return ZONE2AREA.get(z, "未知")

# ===================== 治療日工具(跳過週日 + 休診日) =====================
def is_treat_day(d): return d.weekday()!=6 and d not in HOLIDAYS
def step(d, fwd):
    for _ in range(20):
        d = d + timedelta(days=1) if fwd else d - timedelta(days=1)
        if is_treat_day(d): return d
    return None
def prev_treat(d): return step(d, False)
def next_treat(d): return step(d, True)

# ===================== 解析「目標分頁」→ 每人每日期狀態 =====================
def _cell(df,r,c):
    if r<0 or r>=len(df) or c<0 or c>=df.shape[1]: return ""
    v=df.iat[r,c]; return "" if pd.isna(v) else str(v).strip()

def parse_date(v):
    if v is None: return None
    if isinstance(v,(pd.Timestamp,datetime)):
        try: return v.date()
        except Exception: return None
    s=str(v).strip()
    import re
    m=re.search(r"(\d{4})\D?(\d{1,2})\D(\d{1,2})(?!\d)", s) or re.search(r"(\d{4})\D?(\d{2})(\d{2})(?!\d)", s)
    if m:
        try: return date(int(m.group(1)),int(m.group(2)),int(m.group(3)))
        except Exception: return None
    return None

def parse_sheet(path, sheet):
    df=pd.read_excel(path,sheet_name=sheet,header=None)
    hr=None
    for i in range(len(df)):
        if _cell(df,i,0)=="卡號" and _cell(df,i,1)=="姓名": hr=i; break
    if hr is None: raise RuntimeError(f"分頁「{sheet}」找不到『卡號/姓名』表頭,請確認檔案格式")
    blocks=[c for c in range(df.shape[1]) if _cell(df,hr,c)=="類別"]
    bdates=[]
    for c in blocks:
        d=None
        for rr in range(hr-1,hr-4,-1):
            d=parse_date(df.iat[rr,c]) if 0<=rr<len(df) else None
            if d: break
        bdates.append(d)
    status={}; name_of={}; by_name={}
    for r in range(hr+1,len(df)):
        card=_cell(df,r,0); name=_cell(df,r,1)
        if not card and not name: continue
        if card and name: name_of[card]=name
        if name and card: by_name.setdefault(name,card)
        if not card: continue
        for c,d in zip(blocks,bdates):
            if d is None: continue
            status.setdefault(card,{})[d]=(_cell(df,r,c),_cell(df,r,c+1),_cell(df,r,c+2))
    monday=bdates[0] if bdates and bdates[0] else None
    return status, name_of, by_name, monday

# ===================== 比對名單(卡號優先 / 姓名補位 / 變號自動更新) =====================
def match_members(roster, status, name_of, by_name):
    members=[]; changed=False
    for m in roster:
        card,name=m["card"],m["name"]
        if card and card in status:
            members.append({"card":card,"name":name_of.get(card,name) or name})
        elif name and name in by_name:
            nc=by_name[name]
            warn(f"⚠ 「{name}」卡號變了:名單 {card or '(空)'} → 班表 {nc};已自動更新名單檔")
            m["card"]=nc; changed=True
            members.append({"card":nc,"name":name})
        else:
            warn(f"❌ 名單上的「{name}({card})」這週班表找不到 → 請確認:離職/換人/整週休假?(本週略過)")
    if changed: save_roster(roster)
    return members

# 某人某天 →('白'/'夜', 區域);不可印回 (None,None)
def kind_area(status, card, d, name=""):
    if d is None: return (None,None)
    rec=status.get(card,{}).get(d)
    if not rec: return (None,None)
    cat,shift,zone=rec; s=shift.strip()
    if is_excluded_shift(s): return (None,None)
    a=zone_area(zone)
    if a=="未知" and zone.strip():
        UNKNOWN_BEDS.setdefault(zone.strip(),set()).add((name,d)); return (None,None)
    if a in PRINT_AREAS:
        if s.startswith("D"): return ("白",a)
        if s.startswith("E"): return ("夜",a)
    return (None,None)

# ===================== 公平:歷史累計 =====================
def load_history(skip_week=None):
    cnt={}; rest={}
    for r in _read_rows(F_HISTORY):
        if r.get("週次")==skip_week: continue
        c=r.get("卡號"); st=r.get("狀態")
        if st=="印": cnt[c]=cnt.get(c,0)+1
        elif st=="休": rest[c]=rest.get(c,0)+1
    return cnt, rest

# ===================== 核心排班(以治療日為軸,區域數一般化) =====================
def assign(members, status, window, hist_cnt):
    idx={m["card"]:i for i,m in enumerate(members)}
    cost={m["card"]:hist_cnt.get(m["card"],0) for m in members}
    p1={T:prev_treat(T) for T in window}      # 白班印製日(前1治療日)
    p2={T:prev_treat(p1[T]) if p1[T] else None for T in window}  # 小夜(前2治療日)

    def cands(T, areas):
        """治療日 T、要印 areas 這些區的候選 → {card:(kind,印製日,區,zone,avoid)}"""
        res={}
        for m in members:
            c=m["card"]; nm=m["name"]
            k1,a1=kind_area(status,c,p1[T],nm)
            if k1=="白" and a1 in areas:
                z=status[c][p1[T]][2]; res[c]=("白",p1[T],a1,z, z.strip() in AVOID)
            k2,a2=kind_area(status,c,p2[T],nm)
            if k2=="夜" and a2 in areas and c not in res:
                z=status[c][p2[T]][2]; res[c]=("夜",p2[T],a2,z, z.strip() in AVOID)
        return res

    slots=[(T,A) for T in window for A in PRINT_AREAS]
    same={(T,A):cands(T,{A}) for (T,A) in slots}
    others={A:[x for x in PRINT_AREAS if x!=A] for A in PRINT_AREAS}
    cross={(T,A):cands(T,set(others[A])) for (T,A) in slots}

    assign_slot={}; slot_info={}; member_days={}
    # 每治療日各區「本區可印人數」(Q2:跨區要從人多的區借)
    area_avail={(T,A):len(same[(T,A)]) for (T,A) in slots}
    def used_day(c,T): return T in member_days.get(c,[])
    def order_same(pool):
        # 避開(22-23/29-30)擺後 → 印越少越優先(Q1 誰還沒印) → 已排越少 → 穩定
        return sorted(pool, key=lambda c:(pool[c][4], cost[c], len(member_days.get(c,[])), idx[c]))
    def order_cross(pool, T):
        # Q2:從「組員較多的區」借優先 → 避開擺後 → 印少 → 穩定
        return sorted(pool, key=lambda c:(-area_avail.get((T,pool[c][2]),0), pool[c][4],
                                          cost[c], len(member_days.get(c,[])), idx[c]))

    # 回合1:同區、一人一週一次
    for (T,A) in sorted(slots, key=lambda s: len(same[s])):
        for c in order_same(same[(T,A)]):
            if member_days.get(c) or used_day(c,T): continue
            kind,src,za,z,av=same[(T,A)][c]
            assign_slot[(T,A)]=c; member_days.setdefault(c,[]).append(T)
            slot_info[(T,A)]=(kind,src,False,av); break
    # 回合2:跨區支援(從組員較多的區借)
    for (T,A) in slots:
        if (T,A) in assign_slot: continue
        for c in order_cross(cross[(T,A)], T):
            if member_days.get(c) or used_day(c,T): continue
            kind,src,za,z,av=cross[(T,A)][c]
            assign_slot[(T,A)]=c; member_days.setdefault(c,[]).append(T)
            slot_info[(T,A)]=(kind,src,True,av)
            warn(f"※ {lab(T)} {A} 沒有本區的人 → 從「{za}」借「{name_of_g.get(c,c)}」過來印"); break
    # 回合3:還是空 → 某人印第二次(Q4:盡量間隔久;同日一+二也允許)
    for (T,A) in slots:
        if (T,A) in assign_slot: continue
        merged=dict(cross[(T,A)]); merged.update(same[(T,A)])
        def gap(c):
            ds=member_days.get(c,[]); return min((abs((T-d).days) for d in ds), default=99)
        cand=[c for c in merged if len(member_days.get(c,[]))<2]   # 上限2次,允許同一天
        cand.sort(key=lambda c:(merged[c][4], -gap(c), cost[c], idx[c]))  # 避開後→間隔大優先→印少
        if cand:
            c=cand[0]; g=gap(c); kind,src,za,z,av=merged[c]; isc=(c not in same[(T,A)])
            assign_slot[(T,A)]=c; member_days.setdefault(c,[]).append(T)
            slot_info[(T,A)]=(kind,src,isc,av)
            note="(同日一+二)" if g==0 else ""
            warn(f"※ {lab(T)} {A} 人手不足 → 安排「{name_of_g.get(c,c)}」印第二次{note}(替代方案)")
    # 仍排不出
    for (T,A) in slots:
        if (T,A) not in assign_slot:
            warn(f"❌ {lab(T)} {A} 完全排不出人!請手動處理(名單上會留白標 ❌)")

    able=set()
    for s in slots: able|=set(same[s].keys())
    return slots, assign_slot, slot_info, member_days, able

WD=["週一","週二","週三","週四","週五","週六","週日"]
def lab(d): return f"{WD[d.weekday()]}{d.month}/{d.day}"
name_of_g={}     # 給 assign 內訊息用的姓名表

# ===================== 主流程 =====================
def run():
    path  = sys.argv[1] if len(sys.argv)>1 else None
    sheet = sys.argv[2] if len(sys.argv)>2 else None
    if not path or not os.path.exists(path):
        print("用法:python 排班.py <班表.xls> [分頁名稱]");
        print("（分頁不填=自動用最後一個分頁／最新一週）"); return

    ensure_configs()
    global ZONE2AREA, AVOID, PRINT_AREAS, SHIFT_EXCLUDE, HOLIDAYS, name_of_g
    ZONE2AREA, AVOID, PRINT_AREAS = load_beds()
    SHIFT_EXCLUDE = load_shifts(); HOLIDAYS = load_holidays()
    if not PRINT_AREAS:
        print("❌ 床號分區.csv 裡沒有任何『可印區域』,請檢查設定檔"); return

    try:
        xls=pd.ExcelFile(path)
    except Exception as e:
        print(f"❌ 打不開班表檔:{e}"); return
    if sheet is None:
        sheet=xls.sheet_names[-1]
        for _sn in reversed(xls.sheet_names):
            try:
                _df=pd.read_excel(path,sheet_name=_sn,header=None)
                if any((_cell(_df,_i,0)=='卡號' and _cell(_df,_i,1)=='姓名') for _i in range(len(_df))):
                    sheet=_sn; break
            except Exception: pass
    if sheet not in xls.sheet_names:
        print(f"❌ 找不到分頁「{sheet}」。可用分頁:{xls.sheet_names}"); return

    status, name_of, by_name, monday = parse_sheet(path, sheet)
    name_of_g=name_of
    if not monday:
        print(f"❌ 分頁「{sheet}」抓不到日期,無法決定治療日"); return

    # 名單窗口:從「這週週一的下一個治療日(=週二)」起算 6 個治療日(自動跳週日/連假)
    window=[]; t=next_treat(monday)
    for _ in range(6):
        if t is None: break
        window.append(t); t=next_treat(t)
    skipped=[d for d in (monday+timedelta(days=i) for i in range(1,8)) if d in HOLIDAYS]
    if skipped: warn("ℹ 視為休診跳過:" + "、".join(lab(d) for d in skipped))

    roster=load_roster()
    members=match_members(roster, status, name_of, by_name)
    hist_cnt, hist_rest = load_history(skip_week=sheet)
    slots, assign_slot, slot_info, member_days, able = assign(members, status, window, hist_cnt)
    nm=lambda c: name_of.get(c,c)

    allcards=[m["card"] for m in members]
    printed=set(member_days.keys())
    notprinted=[c for c in allcards if c not in printed]      # 沒印的人=休息(都算公平)
    forced=[c for c in notprinted if c not in able]           # 完全沒可印日(病房/ICU/大夜/休假)

    # ---------- 螢幕 ----------
    print(f"\n■ 檔案:{os.path.basename(path)}　分頁:{sheet}")
    print(f"■ 可印區域:{'、'.join(PRINT_AREAS)}　名單治療日:{lab(window[0])} ~ {lab(window[-1])}")
    print(f"■ 規則:白班印前1治療日、小夜印前2治療日(自動跳週日/連假)\n")
    print("="*16+f" 下週印藥水名單(按治療日) "+"="*16)
    print(f'{"":6}'+"".join(f'{lab(T):<15}' for T in window))
    for A in PRINT_AREAS:
        line=f'{A:<6}'
        for T in window:
            c=assign_slot.get((T,A))
            if not c: line+=f'{"❌排不出":<15}'; continue
            kind,src,isc,av=slot_info[(T,A)]
            tag=f"({src.month}/{src.day}印)"
            if isc: tag+="🔺"
            if len(member_days.get(c,[]))>=2: tag+="雙"
            line+=f'{nm(c)+tag:<15}'
        print(line)
    print("-"*64)
    print("本週休息(都算進公平輪序):", "、".join(nm(c) for c in notprinted) or "無")
    if forced:
        print("  ↳ 其中因病房/ICU/大夜/休假當週不能印:", "、".join(nm(c) for c in forced))

    print("\n【公平累計】(含本週;休=沒印的週數,病房/休假也算)")
    nc=dict(hist_cnt); nr=dict(hist_rest)
    for c in printed:    nc[c]=nc.get(c,0)+1
    for c in notprinted: nr[c]=nr.get(c,0)+1
    for m in sorted(members, key=lambda m:nc.get(m["card"],0)):
        c=m["card"]; mk="印" if c in printed else "休"
        print(f'  {m["name"]:<5} 印{nc.get(c,0):>2} 休{nr.get(c,0):>2}  本週:{mk}')

    if UNKNOWN_BEDS:
        print("\n【⚠ 未知床號(沒在 床號分區.csv 裡,已當不可印,請補設定)】")
        for z,who in UNKNOWN_BEDS.items():
            ppl="、".join(sorted({f"{n}" for n,_ in who if n}))
            print(f"  床號「{z}」  出現於:{ppl}")
    if WARN:
        print("\n【提醒/警告】")
        for w in WARN: print("  "+w)

    # ---------- 輸出 Excel + 文字 ----------
    def celltext(T,A):
        c=assign_slot.get((T,A))
        if not c: return "❌排不出"
        kind,src,isc,av=slot_info[(T,A)]
        t=nm(c)+f"({src.month}/{src.day}印)"
        if isc: t+="🔺"
        if len(member_days.get(c,[]))>=2: t+="雙印"
        return t
    cols={lab(T):[celltext(T,A) for A in PRINT_AREAS] for T in window}
    out_x=os.path.join(OUT_DIR,f"印藥水名單_{sheet}.xlsx")
    try:
        pd.DataFrame(cols, index=PRINT_AREAS).to_excel(out_x)
        _color_excel(out_x, cols)
        print(f"\n✔ Excel:{out_x}")
    except Exception as e:
        warn(f"⚠ 寫 Excel 失敗:{e}")

    lines=[f"📋 下週印藥水名單(治療日 {lab(window[0])}~{lab(window[-1])})",""]
    for T in window:
        parts=[]
        for A in PRINT_AREAS:
            c=assign_slot.get((T,A)); parts.append(f"{A}:{nm(c) if c else '❌排不出'}")
        lines.append(f"{lab(T)}　"+"　".join(parts))
    lines+=["","休息:"+("、".join(nm(c) for c in notprinted) or "無")]
    out_t=os.path.join(OUT_DIR,f"印藥水名單_{sheet}.txt")
    try:
        with open(out_t,"w",encoding="utf-8") as f: f.write("\n".join(lines))
        print(f"✔ 文字(可貼LINE):{out_t}")
    except Exception as e:
        warn(f"⚠ 寫文字檔失敗:{e}")

    # ---------- 雲端貼上版(印日期,區,姓名)→ 給 Google 表 / LINE 提醒 ----------
    out_c=os.path.join(OUT_DIR,f"雲端貼上_{sheet}.csv")
    try:
        with open(out_c,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.writer(f); w.writerow(["印日期","區","姓名"])
            for T in window:
                for A in PRINT_AREAS:
                    c=assign_slot.get((T,A))
                    if not c: continue
                    src=slot_info[(T,A)][1]
                    w.writerow([f"{src.year}-{src.month:02d}-{src.day:02d}", A, nm(c)])
        print(f"\u2714 雲端貼上版：{out_c}")
    except Exception as e:
        warn(f"⚠ 寫雲端貼上版失敗：{e}")

    # ---------- 寫回歷史(同一週覆蓋,不重複累計) ----------
    rows=[r for r in _read_rows(F_HISTORY) if r.get("週次")!=sheet]
    for m in members:
        c=m["card"]; st="印" if c in printed else "休"
        days=";".join(lab(T) for T in member_days.get(c,[]))
        rows.append({"週次":sheet,"卡號":c,"姓名":m["name"],"狀態":st,"治療日":days})
    try:
        with open(F_HISTORY,"w",newline="",encoding="utf-8-sig") as f:
            w=csv.DictWriter(f,fieldnames=["週次","卡號","姓名","狀態","治療日"]); w.writeheader()
            for r in rows: w.writerow(r)
    except Exception as e:
        print(f"⚠ 寫排班紀錄失敗(不影響本次名單):{e}")

def _color_excel(path, cols):
    """把『排不出』標紅、跨區/雙印標橙(O1)。失敗就算了,不影響名單。"""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        red=PatternFill("solid",fgColor="FFC7CE"); orange=PatternFill("solid",fgColor="FFE5B4")
        wb=load_workbook(path); ws=wb.active
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                v=str(cell.value or "")
                if "排不出" in v: cell.fill=red
                elif "🔺" in v or "雙印" in v: cell.fill=orange
        wb.save(path)
    except Exception:
        pass

def main():
    try:
        run()
    except Exception as e:
        print("\n❌ 程式遇到沒預期到的狀況,但我幫你攔下來了(不會弄壞檔案)。")
        print("   錯誤訊息:", e)
        print("   detail：")
        traceback.print_exc()

if __name__=="__main__":
    main()
